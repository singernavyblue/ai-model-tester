#!/usr/bin/env python3
"""
清洗测试结果 Excel：清除 output_content 中的题号前缀和语言指令，非中文翻译为中文。

用法:
    python clean_results.py --input <xlsx文件> [--translate-key <DeepSeek API key>]

示例:
    python clean_results.py --input 7.2汇总.xlsx
    python clean_results.py --input 7.2汇总.xlsx --translate-key sk-xxx
"""

import re
import sys
import argparse
import json
import time
import urllib.request
from pathlib import Path
from openpyxl import load_workbook


def clean_prefix(text: str) -> str:
    """清除 output_content 中的题号前缀和语言指令。"""
    if not text:
        return text
    text = str(text)
    # 去掉开头的 [1.1] 或 [1] 题号
    text = re.sub(r'^\[\d+(?:\.\d+)?\]\s*', '', text)
    # 去掉【重要指令】... 整段
    text = re.sub(
        r'\n*【重要指令】你必须只使用中文回答下面的问题。不要使用任何其他语言，无论问题是什么语言，你的回答都必须是中文。[。]?\s*',
        '', text
    )
    # 去掉旧版指令
    text = re.sub(r'\n*请用中文回答以下问题：\s*', '', text)
    # 去掉单独的 "今天" 或 "今天，" 开头
    text = re.sub(r'^今天[，,]?\s*', '', text)
    return text.strip()


def is_chinese(text: str) -> bool:
    """判断文本是否主要为中文。"""
    if not text:
        return True
    cn = sum(1 for c in str(text) if '一' <= c <= '鿿')
    total = max(len(str(text).replace(' ', '').replace('\n', '')), 1)
    return cn > total * 0.3


def translate_text(text: str, api_key: str) -> str:
    """使用 DeepSeek 翻译为中文。"""
    prompt = f"请将以下内容翻译成中文，只输出翻译结果，不要添加任何解释：\n\n{text[:8000]}"
    try:
        url = "https://api.deepseek.com/v1/chat/completions"
        body = json.dumps({
            "model": "deepseek-chat",
            "max_tokens": 4096,
            "messages": [{"role": "user", "content": prompt}],
        }).encode()
        req = urllib.request.Request(url, data=body, headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        })
        with urllib.request.urlopen(req, timeout=180) as resp:
            data = json.loads(resp.read().decode())
        return data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        print(f"  翻译失败: {e}")
        return text


def clean_file(filepath: str, translate_key: str = None):
    """清洗单个 xlsx 文件。"""
    print(f"📖 打开: {filepath}")
    wb = load_workbook(filepath)
    ws = wb.active
    hdrs = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    out_col = hdrs.get('output_content')
    note_col = hdrs.get('note')
    inp_col = hdrs.get('input_content')

    if not out_col:
        print("❌ 未找到 output_content 列")
        return

    # Step 1: 清除前缀
    cleaned_prefix = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[out_col - 1]
        if not cell.value:
            continue
        new_text = clean_prefix(str(cell.value))
        if new_text != cell.value:
            cell.value = new_text
            cleaned_prefix += 1
    print(f"  清除前缀: {cleaned_prefix} 个单元格")

    # Also clean input_content
    if inp_col:
        cleaned_inp = 0
        for row in ws.iter_rows(min_row=2):
            cell = row[inp_col - 1]
            if not cell.value:
                continue
            new_text = clean_prefix(str(cell.value))
            if new_text != cell.value:
                cell.value = new_text
                cleaned_inp += 1
        print(f"  清除 input 前缀: {cleaned_inp} 个单元格")

    # Step 2: 翻译非中文
    if translate_key:
        non_cn_rows = []
        for row in ws.iter_rows(min_row=2):
            text = str(row[out_col - 1].value) if row[out_col - 1].value else ''
            if text and not is_chinese(text):
                non_cn_rows.append(row)

        if non_cn_rows:
            print(f"  需翻译: {len(non_cn_rows)} 行")
            translated = 0
            for i, row in enumerate(non_cn_rows):
                row_num = row[0].row
                text = str(row[out_col - 1].value)
                print(f"    [{i+1}/{len(non_cn_rows)}] 行{row_num} ({len(text)}字)...",
                      end=" ", flush=True)
                result = translate_text(text, translate_key)
                if result != text:
                    row[out_col - 1].value = result
                    translated += 1
                    # 标记到 note
                    if note_col:
                        nc = row[note_col - 1]
                        old = str(nc.value) if nc.value else ''
                        nc.value = (old + '；已翻译为中文' if old else '已翻译为中文').strip('；')
                    print('✅')
                else:
                    print('❌')
                time.sleep(0.5)
            print(f"  翻译完成: {translated}/{len(non_cn_rows)}")
        else:
            print(f"  无需翻译，所有 output 已是中文")
    else:
        # 仅检测不翻译
        non_cn = sum(1 for row in ws.iter_rows(min_row=2)
                     if row[out_col - 1].value and not is_chinese(str(row[out_col - 1].value)))
        if non_cn > 0:
            print(f"  ⚠️ 检测到 {non_cn} 行非中文，使用 --translate-key 进行翻译")

    wb.save(filepath)
    print(f"✅ 保存: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="清洗测试结果 Excel")
    parser.add_argument("--input", "-i", required=True, help="要清洗的 xlsx 文件路径")
    parser.add_argument("--translate-key", "-k", default=None,
                        help="DeepSeek API key 用于翻译非中文内容")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"❌ 文件不存在: {args.input}")
        sys.exit(1)

    clean_file(str(path), args.translate_key)


if __name__ == "__main__":
    main()
