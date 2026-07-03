#!/usr/bin/env python3
"""
合并多个测试结果 Excel 文件为一个文件。

用法:
    python merge_results.py --input <文件夹路径> [--output <输出文件>] [--pattern <文件名匹配>]

示例:
    # 合并测试结果文件夹下所有 xlsx
    python merge_results.py --input 测试结果/

    # 只合并 7.2 相关文件
    python merge_results.py --input 测试结果/ --pattern "7.2*网页测试*"

    # 指定输出文件名
    python merge_results.py --input 测试结果/ --output 7.2汇总.xlsx
"""

import os
import sys
import argparse
import glob
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage


def merge_files(file_paths: list[str], output_path: str):
    """合并多个测试结果 xlsx 文件。"""
    if not file_paths:
        print("❌ 没有找到要合并的文件")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试详细结果"

    # Styles
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_font = Font(name="微软雅黑", size=10)
    c_align = Alignment(vertical="top", wrap_text=True)
    c_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    border = Border(left=Side("thin", "D9D9D9"), right=Side("thin", "D9D9D9"),
                    top=Side("thin", "D9D9D9"), bottom=Side("thin", "D9D9D9"))

    headers = ["record_id", "monitoring_batch", "date", "time_point", "model",
               "model_family", "language", "input_content", "output_content",
               "is_anomaly", "anomaly_reason", "note", "screenshot"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

    record_id = 1
    new_row = 2
    total_imgs = 0

    for f in sorted(file_paths):
        print(f"  合并: {os.path.basename(f)} ...", end=" ")
        wb_src = load_workbook(f)
        ws_src = wb_src.active
        count = ws_src.max_row - 1
        src_hdrs = {ws_src.cell(row=1, column=c).value: c for c in range(1, ws_src.max_column + 1)}

        for src_row in ws_src.iter_rows(min_row=2, values_only=True):
            vals = [
                record_id,
                str(src_row[src_hdrs.get('monitoring_batch', 2)-1] or '') if src_hdrs.get('monitoring_batch') else '',
                str(src_row[src_hdrs.get('date', 3)-1] or '') if src_hdrs.get('date') else '',
                str(src_row[src_hdrs.get('time_point', 4)-1] or '') if src_hdrs.get('time_point') else '',
                str(src_row[src_hdrs.get('model', 5)-1] or '') if src_hdrs.get('model') else '',
                str(src_row[src_hdrs.get('model_family', 6)-1] or '') if src_hdrs.get('model_family') else '',
                str(src_row[src_hdrs.get('language', 7)-1] or '') if src_hdrs.get('language') else '',
                src_row[src_hdrs.get('input_content', 8)-1] if src_hdrs.get('input_content') else '',
                src_row[src_hdrs.get('output_content', 9)-1] if src_hdrs.get('output_content') else '',
                str(src_row[src_hdrs.get('is_anomaly', 10)-1] or '') if src_hdrs.get('is_anomaly') else '',
                str(src_row[src_hdrs.get('anomaly_reason', 11)-1] or '') if src_hdrs.get('anomaly_reason') else '',
                str(src_row[src_hdrs.get('note', 12)-1] or '') if src_hdrs.get('note') else '',
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=new_row, column=c, value=v)
                cell.font = c_font; cell.border = border
                cell.alignment = c_center if c in (1, 3, 4, 5, 6, 7, 10) else c_align

            # Copy screenshot
            ss_col = src_hdrs.get('screenshot')
            if ss_col:
                ss_name = str(src_row[ss_col - 1]) if len(src_row) >= ss_col and src_row[ss_col - 1] else ''
                if ss_name and ss_name.endswith('.png'):
                    ss_path = os.path.join(os.path.dirname(f), '截图', ss_name)
                    if os.path.exists(ss_path):
                        try:
                            img = XlImage(ss_path)
                            ratio = 320 / img.width
                            img.width = 320
                            img.height = int(img.height * ratio)
                            ws.add_image(img, f"M{new_row}")
                            ws.cell(row=new_row, column=13, value=ss_name).font = c_font
                            ws.cell(row=new_row, column=13).alignment = c_center
                            ws.cell(row=new_row, column=13).border = border
                            total_imgs += 1
                        except Exception:
                            pass
            record_id += 1
            new_row += 1
        print(f"{count} 行")

    # Row heights
    for i in range(2, new_row):
        ss = ws.cell(row=i, column=13).value
        ws.row_dimensions[i].height = 200 if (ss and str(ss).endswith('.png')) else 16
    ws.row_dimensions[1].height = 16

    # Column widths
    for col, w in {1: 8, 2: 14, 3: 14, 4: 12, 5: 18, 6: 18, 7: 10,
                   8: 50, 9: 60, 10: 12, 11: 28, 12: 20, 13: 38}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # Sheet 2: 汇总统计
    ws2 = wb.create_sheet("汇总统计")
    for c, h in enumerate(["模型", "总题数", "成功", "失败", "异常数"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

    model_stats = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        m = str(row[4])
        model_stats.setdefault(m, {"total": 0, "anomaly": 0})
        model_stats[m]["total"] += 1
        if row[9] == "TRUE":
            model_stats[m]["anomaly"] += 1

    for i, (m, s) in enumerate(sorted(model_stats.items()), 2):
        for c, v in enumerate([m, s["total"], s["total"], 0, s["anomaly"]], 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = c_font; cell.alignment = c_center; cell.border = border
    for row in ws2.iter_rows(min_row=1, max_row=len(model_stats) + 1):
        ws2.row_dimensions[row[0].row].height = 16
    for col, w in {1: 24, 2: 10, 3: 8, 4: 8, 5: 10}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet 3: 按分类汇总
    ws3 = wb.create_sheet("按分类汇总")
    all_models = sorted(set(str(r[4]) for r in ws.iter_rows(min_row=2, values_only=True)))
    for c, h in enumerate(["语言"] + [f"{m}\n(异常/总数)" for m in all_models], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

    lang_stats = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        lang = str(row[6])
        model = str(row[4])
        lang_stats.setdefault(lang, {}).setdefault(model, {"total": 0, "anomaly": 0})
        lang_stats[lang][model]["total"] += 1
        if row[9] == "TRUE":
            lang_stats[lang][model]["anomaly"] += 1

    for i, (lang, md) in enumerate(sorted(lang_stats.items()), 2):
        ws3.cell(row=i, column=1, value=lang).font = c_font
        ws3.cell(row=i, column=1).alignment = c_align
        ws3.cell(row=i, column=1).border = border
        for j, mn in enumerate(all_models, 2):
            d = md.get(mn, {"total": 0, "anomaly": 0})
            cell = ws3.cell(row=i, column=j, value=f"{d['anomaly']}/{d['total']}")
            cell.font = c_font; cell.alignment = c_center; cell.border = border

    ws3.column_dimensions["A"].width = 14
    for j in range(2, len(all_models) + 2):
        ws3.column_dimensions[get_column_letter(j)].width = 18
    for row in ws3.iter_rows(min_row=1, max_row=len(lang_stats) + 1):
        ws3.row_dimensions[row[0].row].height = 16
    ws3.freeze_panes = "B2"

    wb.save(output_path)
    print(f"\n✅ 合并完成: {output_path}")
    print(f"   总行数: {record_id - 1}")
    print(f"   文件数: {len(file_paths)}")
    print(f"   嵌入截图: {total_imgs}")
    return record_id - 1


def main():
    parser = argparse.ArgumentParser(description="合并多个测试结果 Excel 文件")
    parser.add_argument("--input", "-i", required=True, help="输入文件夹路径")
    parser.add_argument("--output", "-o", default=None, help="输出文件路径")
    parser.add_argument("--pattern", "-p", default="网页测试结果_*.xlsx",
                        help="文件匹配模式（默认: 网页测试结果_*.xlsx）")
    args = parser.parse_args()

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        print(f"❌ 文件夹不存在: {args.input}")
        sys.exit(1)

    files = sorted(glob.glob(str(input_dir / args.pattern)))
    if not files:
        print(f"❌ 没有匹配的文件: {args.pattern}")
        sys.exit(1)

    if args.output is None:
        args.output = str(input_dir / f"合并汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    print(f"文件夹: {args.input}")
    print(f"匹配: {args.pattern}")
    print(f"找到: {len(files)} 个文件")
    print(f"输出: {args.output}")
    print()

    merge_files(files, args.output)


if __name__ == "__main__":
    main()
