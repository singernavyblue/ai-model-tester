#!/usr/bin/env python3
"""
检测汇总 Excel 中的空白回答和网络错误，只重测这些问题行。

用法:
    python retry_blanks.py --input 汇总.xlsx --services meta \
        --judge-key sk-xxx [--no-screenshot]

工作流程:
    1. 扫描 Excel，找到 output_content 为空或 anomaly_reason 包含网络错误的行
    2. 生成临时 docx 仅包含这些问题
    3. 用网页自动化逐个重测
    4. 将新结果写回原 Excel 对应行
    5. 清理临时文件
"""

import os, sys, re, json, time, argparse, tempfile, shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage


def find_blank_rows(filepath: str) -> list[dict]:
    """扫描 Excel，返回空白/错误行列表。"""
    wb = load_workbook(filepath)
    ws = wb.active
    hdrs = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    blanks = []
    for row in ws.iter_rows(min_row=2):
        out_col = hdrs.get('output_content')
        note_col = hdrs.get('note')
        inp_col = hdrs.get('input_content')
        lang_col = hdrs.get('language')
        model_col = hdrs.get('model')

        out_val = str(row[out_col - 1].value) if out_col and row[out_col - 1].value else ''
        note_val = str(row[note_col - 1].value) if note_col and row[note_col - 1].value else ''
        inp_val = str(row[inp_col - 1].value) if inp_col and row[inp_col - 1].value else ''
        lang_val = str(row[lang_col - 1].value) if lang_col and len(row) >= lang_col and row[lang_col - 1].value else ''
        model_val = str(row[model_col - 1].value) if model_col and len(row) >= model_col and row[model_col - 1].value else ''

        is_blank = not out_val or len(out_val) < 50
        is_error = any(kw in note_val for kw in ['Timeout', '无法访问', 'connection', 'Connection', 'ERR_'])

        if is_blank or is_error:
            blanks.append({
                'row_num': row[0].row,
                'language': lang_val,
                'model': model_val,
                'input_content': inp_val,
                'is_blank': is_blank,
                'is_error': is_error,
            })

    wb.close()
    return blanks


def create_temp_docx(questions: list[dict], output_path: str):
    """创建仅包含问题行的临时 docx。"""
    from zipfile import ZipFile, ZIP_DEFLATED
    from xml.etree.ElementTree import Element, SubElement, tostring

    # 构建 document.xml
    ns = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
    doc = Element(f'{{{ns}}}document')
    body = SubElement(doc, f'{{{ns}}}body')

    for q in questions:
        p = SubElement(body, f'{{{ns}}}p')
        r = SubElement(p, f'{{{ns}}}r')
        t = SubElement(r, f'{{{ns}}}t')
        t.text = f"{q['question_num']}{q['question_text']}"
        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    doc_xml = tostring(doc, encoding='unicode')

    # 构建最小 docx zip
    with ZipFile(output_path, 'w', ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>''')
        zf.writestr('_rels/.rels', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>''')
        zf.writestr('word/document.xml', doc_xml)


def retest_and_update(xlsx_path: str, blanks: list[dict],
                      services: list[str], judge_key: str = None,
                      screenshot: bool = True, delay: float = 3.0):
    """对空白行逐题重测，写回 Excel。"""
    # Import here to avoid circular deps
    sys.path.insert(0, str(Path(__file__).parent))
    from run_browser_test import run_browser_test, parse_docx

    # 构建问题列表（保留行号映射）
    questions = []
    for i, b in enumerate(blanks):
        # 从原始 input 提取问题文本
        q_text = b['input_content'].strip()
        questions.append({
            'question_num': str(i + 1),
            'question_text': q_text,
            'row_num': b['row_num'],
            'doc_lang': b['language'],
        })

    if not questions:
        print("没有需要重测的行")
        return

    print(f"\n🔄 重测 {len(questions)} 道题目...\n")

    # 创建临时 docx
    tmpdir = tempfile.mkdtemp()
    tmp_docx = os.path.join(tmpdir, 'retry.docx')
    create_temp_docx(questions, tmp_docx)

    # 用网页模式测试
    results = run_browser_test(
        services=services,
        questions=questions,
        output_path=os.path.join(tmpdir, 'temp_result.xlsx'),
        headless=False,
        question_delay=delay,
        answer_lang='zh',
        doc_lang='',
        question_limit=0,
        monitoring_batch='retry',
        retry_rounds=0,
        enable_screenshot=screenshot,
    )

    # 解析结果
    wb = load_workbook(xlsx_path)
    ws = wb.active
    hdrs = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    out_col = hdrs.get('output_content')
    note_col = hdrs.get('note')

    updated = 0
    for r in results:
        row_num = r.get('_row_num') if '_row_num' in r else None
        # results from run_browser_test contain question_text, find matching row
        if row_num is None:
            # Match by question_text
            q_text = r.get('original_question', r.get('question_text', ''))
            for b in blanks:
                if b['input_content'].strip() == q_text.strip():
                    row_num = b['row_num']
                    break
        if row_num is None:
            continue

        new_out = r.get('model_response', '')
        if new_out and len(new_out) > 50:
            ws.cell(row=row_num, column=out_col).value = new_out
            if note_col:
                old_note = str(ws.cell(row=row_num, column=note_col).value or '')
                ws.cell(row=row_num, column=note_col).value = (old_note + '；已重测').strip('；')
            updated += 1
            print(f"  ✅ 行{row_num} 已更新")

    wb.save(xlsx_path)
    wb.close()

    # 清理
    shutil.rmtree(tmpdir, ignore_errors=True)
    print(f"\n重测完成: 更新 {updated}/{len(questions)} 行")


def main():
    parser = argparse.ArgumentParser(description="检测空白回答并重测")
    parser.add_argument("--input", "-i", required=True, help="汇总 Excel 文件路径")
    parser.add_argument("--services", "-s", default="meta", help="测试服务（默认 meta）")
    parser.add_argument("--judge-key", default=None, help="审核 API key")
    parser.add_argument("--no-screenshot", action="store_true", help="关闭截图")
    parser.add_argument("--delay", "-d", type=float, default=3.0, help="题间延迟秒数")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"❌ 文件不存在: {args.input}")
        sys.exit(1)

    print(f"📖 扫描: {args.input}")
    blanks = find_blank_rows(str(path))

    if not blanks:
        print("✅ 没有空白行，无需重测")
        return

    print(f"⚠️  发现 {len(blanks)} 行需要重测:")
    for b in blanks:
        tag = "空白" if b['is_blank'] else "网络错误"
        print(f"    行{b['row_num']} [{b['language']}] {tag} - {b['input_content'][:50]}...")

    services = [s.strip() for s in args.services.split(",")]
    retest_and_update(
        str(path), blanks, services,
        judge_key=args.judge_key or os.environ.get("DEEPSEEK_API_KEY", ""),
        screenshot=not args.no_screenshot,
        delay=args.delay,
    )

    # 检测是否还有残留空白
    remaining = find_blank_rows(str(path))
    if remaining:
        print(f"\n⚠️  仍有 {len(remaining)} 行空白，可能存在网络问题需稍后重试")
    else:
        print(f"\n✅ 所有行已填充完毕")


if __name__ == "__main__":
    main()
